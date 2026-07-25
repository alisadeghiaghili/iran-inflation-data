"""Excel exporter for Iran inflation data."""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..config import OUTPUT_DIR


def _style_worksheet(ws, headers: list[str]) -> None:
    """Apply formatting to a worksheet."""
    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)

    # Freeze header row
    ws.freeze_panes = "A2"


def export_excel(monthly_df: pd.DataFrame | None, yearly_df: pd.DataFrame | None) -> None:
    """Export data to a multi-sheet Excel workbook.

    Sheets:
        - Monthly Data
        - Yearly Data
        - Summary
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    excel_path = os.path.join(OUTPUT_DIR, "iran_inflation.xlsx")

    has_monthly = monthly_df is not None and not monthly_df.empty
    has_yearly = yearly_df is not None and not yearly_df.empty

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Monthly sheet
        if has_monthly:
            monthly_df.to_excel(writer, sheet_name="Monthly Data", index=False, startrow=1)
            ws = writer.sheets["Monthly Data"]
            _style_worksheet(ws, list(monthly_df.columns))

        # Yearly sheet
        if has_yearly:
            yearly_df.to_excel(writer, sheet_name="Yearly Data", index=False, startrow=1)
            ws = writer.sheets["Yearly Data"]
            _style_worksheet(ws, list(yearly_df.columns))

        # Summary sheet
        ws_summary = writer.book.create_sheet("Summary")
        summary_data = [
            ["Iran Inflation Data - Summary", ""],
            ["", ""],
            ["Data Source", "Records"],
        ]

        if has_monthly:
            summary_data.append(["Monthly Data (CBI/IMF)", len(monthly_df)])
            date_range = f"{monthly_df['date'].min().strftime('%Y-%m')} to {monthly_df['date'].max().strftime('%Y-%m')}"
            summary_data.append(["  Date Range", date_range])
        else:
            summary_data.append(["Monthly Data (CBI/IMF)", "No data available"])

        if has_yearly:
            summary_data.append(["Yearly Data (World Bank)", len(yearly_df)])
            summary_data.append(["  Year Range", f"{int(yearly_df['year'].min())} to {int(yearly_df['year'].max())}"])
        else:
            summary_data.append(["Yearly Data (World Bank)", "No data available"])

        summary_data.extend([
            ["", ""],
            ["Columns in Yearly Data", ""],
            ["  year", "Year of observation"],
            ["  cpi_index", "Consumer Price Index (2010=100)"],
            ["  inflation_yoy_pct", "Year-over-year inflation rate (%)"],
            ["  food_inflation_pct", "Food inflation rate (%)"],
            ["  source", "Data source (World Bank)"],
            ["", ""],
            ["Columns in Monthly Data", ""],
            ["  date", "Date (YYYY-MM-DD)"],
            ["  year", "Year"],
            ["  month", "Month (1-12)"],
            ["  cpi_index", "Consumer Price Index value"],
            ["  source", "Data source (CBI or IMF)"],
        ])

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_summary.cell(row=row_idx, column=col_idx, value=value)

        # Style summary title
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws_summary.column_dimensions["A"].width = 35
        ws_summary.column_dimensions["B"].width = 40

    print(f"[Excel] Exported data to {excel_path}")
    if has_monthly:
        print(f"  Monthly Data: {len(monthly_df)} rows")
    if has_yearly:
        print(f"  Yearly Data: {len(yearly_df)} rows")
